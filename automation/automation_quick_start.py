#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动化脚本快速入门教程
=====================
适合初中生学习的自动化脚本教程

作者: Python学习小组
预计学习时间: 约2小时

运行方式:
    python automation_quick_start.py

学习建议:
    1. 按顺序学习每一节
    2. 动手修改代码尝试不同效果
    3. 完成每节后的练习题
"""

import os
import shutil
import time
import random
from datetime import datetime

# ============================================================
#                    第1节：什么是自动化？
# ============================================================

def section_1_what_is_automation():
    """
    第1节：什么是自动化？

    【概念解释】
    自动化就是让计算机自动完成重复性的工作。
    想象一下：
    - 你需要把100个文件按类型分类 → 手动需要1小时，自动化只需几秒
    - 你需要给50个文件重命名 → 手动容易出错，自动化又快又准
    - 你需要每天备份数据 → 自动化可以在你睡觉时完成

    【生活中的自动化例子】
    1. 手机闹钟：每天固定时间响铃
    2. 自动回复：收到邮件自动回复
    3. 批量处理：一次性修改多张图片大小
    4. 数据备份：定期自动备份重要文件

    【为什么要学自动化？】
    1. 省时间：把重复工作交给电脑
    2. 少出错：电脑不会疲劳，不会手滑
    3. 更专业：提高工作效率
    4. 有成就感：看着程序自动完成任务很酷！

    【Python自动化的优势】
    1. 简单易学：代码像英语一样容易理解
    2. 功能强大：可以操作文件、Excel、网页等
    3. 免费开源：不需要花钱购买
    4. 社区活跃：有问题很容易找到答案
    """
    print("=" * 60)
    print("第1节：什么是自动化？")
    print("=" * 60)
    print("""
【自动化就是让电脑帮你做重复的工作！】

生活中的例子：
  📱 手机闹钟 - 每天固定时间自动响铃
  📧 自动回复 - 收到邮件自动回复
  🔄 批量处理 - 一次性处理多个文件
  💾 自动备份 - 定期备份重要数据

为什么要学？
  ⏰ 省时间 - 几秒完成几小时的工作
  ✅ 少出错 - 电脑不会疲劳
  🚀 更专业 - 提高工作效率
  😎 有成就感 - 看着程序自动运行很酷

Python是自动化最好的选择：
  ✨ 简单易学 - 代码像英语一样
  ✨ 功能强大 - 文件、Excel、网页都能搞定
  ✨ 免费开源 - 不花一分钱
    """)

    # 演示：一个简单的自动化例子
    print("【小演示】自动化 vs 手动")
    print("-" * 40)

    # 模拟手动操作
    print("手动操作：一个一个打印1到5")
    print("  print(1)")
    print("  print(2)")
    print("  print(3)")
    print("  print(4)")
    print("  print(5)")
    print("  结果: ", end="")
    print(1, 2, 3, 4, 5)

    print()

    # 自动化操作
    print("自动化操作：用循环一次搞定")
    print("  for i in range(1, 6):")
    print("      print(i)")
    print("  结果: ", end="")
    for i in range(1, 6):
        print(i, end=" ")
    print("\n")

    print("💡 当数字变成100个、1000个时，自动化的优势就更明显了！")
    print("=" * 60)
    print()


# ============================================================
#                  第2节：文件操作自动化
# ============================================================

def section_2_file_operations():
    """
    第2节：文件操作自动化

    【学习目标】
    学会用Python创建、复制、移动、删除文件和文件夹

    【核心知识点】
    1. os模块：操作系统相关的功能
    2. shutil模块：高级文件操作

    【重要函数说明】

    os.mkdir(路径)
        功能：创建文件夹
        参数：路径 - 要创建的文件夹路径
        示例：os.mkdir("test_folder")

    os.path.exists(路径)
        功能：检查文件或文件夹是否存在
        参数：路径 - 要检查的路径
        返回：True（存在）或 False（不存在）
        示例：os.path.exists("test.txt")

    shutil.copy(源文件, 目标位置)
        功能：复制文件
        参数：源文件 - 要复制的文件路径
             目标位置 - 复制到哪里
        示例：shutil.copy("a.txt", "b.txt")

    shutil.move(源文件, 目标位置)
        功能：移动文件（也可以重命名）
        参数：源文件 - 要移动的文件路径
             目标位置 - 移动到哪里
        示例：shutil.move("a.txt", "folder/a.txt")

    shutil.rmtree(文件夹路径)
        功能：删除整个文件夹（包括里面的所有内容）
        参数：文件夹路径 - 要删除的文件夹
        ⚠️ 警告：删除后无法恢复，谨慎使用！
        示例：shutil.rmtree("test_folder")
    """
    print("=" * 60)
    print("第2节：文件操作自动化")
    print("=" * 60)

    # 设置练习目录
    practice_dir = "test_files/practice_section2"

    print("\n【核心模块介绍】")
    print("-" * 40)
    print("os模块     - 操作系统功能，如创建文件夹、检查路径")
    print("shutil模块 - 高级文件操作，如复制、移动、删除")

    print("\n【动手实践】让我们创建一些文件和文件夹")
    print("-" * 40)

    # 1. 创建文件夹
    print("\n1️⃣  创建文件夹")
    if not os.path.exists(practice_dir):
        os.makedirs(practice_dir)
        print(f"   ✅ 创建文件夹：{practice_dir}")
    else:
        print(f"   📁 文件夹已存在：{practice_dir}")

    # 2. 创建文件
    print("\n2️⃣  创建文件")
    test_file = os.path.join(practice_dir, "hello.txt")
    with open(test_file, "w", encoding="utf-8") as f:
        f.write("你好，这是自动化创建的文件！\n")
        f.write("Python自动化真强大！\n")
    print(f"   ✅ 创建文件：{test_file}")

    # 3. 复制文件
    print("\n3️⃣  复制文件")
    copy_file = os.path.join(practice_dir, "hello_copy.txt")
    shutil.copy(test_file, copy_file)
    print(f"   ✅ 复制到：{copy_file}")

    # 4. 移动/重命名文件
    print("\n4️⃣  重命名文件")
    renamed_file = os.path.join(practice_dir, "hello_renamed.txt")
    shutil.move(copy_file, renamed_file)
    print(f"   ✅ 重命名为：{renamed_file}")

    # 5. 列出文件夹内容
    print("\n5️⃣  列出文件夹内容")
    files = os.listdir(practice_dir)
    for f in files:
        print(f"   📄 {f}")

    # 6. 检查文件信息
    print("\n6️⃣  检查文件信息")
    file_size = os.path.getsize(test_file)
    print(f"   文件大小：{file_size} 字节")

    # 清理演示文件
    print("\n7️⃣  清理演示文件")
    shutil.rmtree(practice_dir)
    print(f"   🗑️  已删除：{practice_dir}")

    print("\n【代码总结】")
    print("-" * 40)
    print("""
常用文件操作代码：

# 创建文件夹
os.makedirs("文件夹路径")

# 检查是否存在
os.path.exists("路径")  # 返回 True/False

# 复制文件
shutil.copy("源文件", "目标位置")

# 移动/重命名
shutil.move("源", "目标")

# 列出文件夹内容
os.listdir("文件夹路径")

# 删除文件夹（慎用！）
shutil.rmtree("文件夹路径")
    """)

    print("=" * 60)
    print()


# ============================================================
#                  第3节：批量重命名
# ============================================================

def section_3_batch_rename():
    """
    第3节：批量重命名

    【学习目标】
    学会用Python批量修改文件名

    【应用场景】
    1. 把"照片(1).jpg"改成"2024_01_01_001.jpg"
    2. 把所有".txt"文件改成".md"文件
    3. 给文件名添加前缀或后缀
    4. 统一文件名格式

    【核心知识点】
    1. os.rename()：重命名单个文件
    2. os.listdir()：获取文件列表
    3. 字符串操作：split(), join(), format()

    【重要函数说明】

    os.rename(旧名称, 新名称)
        功能：重命名文件或文件夹
        参数：旧名称 - 原来的文件名
             新名称 - 新的文件名
        示例：os.rename("old.txt", "new.txt")

    os.path.splitext(文件名)
        功能：分离文件名和扩展名
        参数：文件名 - 包含扩展名的文件名
        返回：(文件名, 扩展名) 元组
        示例：os.path.splitext("photo.jpg") → ("photo", ".jpg")
    """
    print("=" * 60)
    print("第3节：批量重命名")
    print("=" * 60)

    # 创建演示文件
    demo_dir = "test_files/rename_demo"
    os.makedirs(demo_dir, exist_ok=True)

    print("\n【准备工作】创建演示文件...")
    for i in range(1, 6):
        with open(os.path.join(demo_dir, f"photo({i}).jpg"), "w") as f:
            f.write(f"This is photo {i}")

    print("原始文件：")
    for f in sorted(os.listdir(demo_dir)):
        print(f"  📄 {f}")

    # 演示1：添加前缀
    print("\n【示例1】添加日期前缀")
    print("-" * 40)

    def add_prefix(folder, prefix):
        """
        给文件夹中的所有文件添加前缀

        参数：
            folder: 文件夹路径
            prefix: 要添加的前缀

        示例调用：
            add_prefix("photos", "2024_")
        """
        for filename in os.listdir(folder):
            old_path = os.path.join(folder, filename)
            # 跳过文件夹
            if os.path.isfile(old_path):
                new_name = prefix + filename
                new_path = os.path.join(folder, new_name)
                os.rename(old_path, new_path)
                print(f"  {filename} → {new_name}")

    add_prefix(demo_dir, "vacation_")

    # 演示2：修改扩展名
    print("\n【示例2】批量修改扩展名")
    print("-" * 40)

    def change_extension(folder, old_ext, new_ext):
        """
        批量修改文件扩展名

        参数：
            folder: 文件夹路径
            old_ext: 原扩展名（如".jpg"）
            new_ext: 新扩展名（如".png"）

        示例调用：
            change_extension("photos", ".jpg", ".png")
        """
        for filename in os.listdir(folder):
            if filename.endswith(old_ext):
                name = os.path.splitext(filename)[0]
                new_name = name + new_ext
                old_path = os.path.join(folder, filename)
                new_path = os.path.join(folder, new_name)
                os.rename(old_path, new_path)
                print(f"  {filename} → {new_name}")

    change_extension(demo_dir, ".jpg", ".png")

    # 演示3：编号格式化
    print("\n【示例3】统一编号格式")
    print("-" * 40)

    def format_numbers(folder, prefix, digits=3):
        """
        统一文件编号格式

        参数：
            folder: 文件夹路径
            prefix: 新文件名前缀
            digits: 编号位数（如3表示001, 002...）

        示例调用：
            format_numbers("photos", "photo_", 3)
            # 结果：photo_001.png, photo_002.png...
        """
        files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
        files.sort()

        for i, filename in enumerate(files, 1):
            ext = os.path.splitext(filename)[1]
            new_name = f"{prefix}{str(i).zfill(digits)}{ext}"
            old_path = os.path.join(folder, filename)
            new_path = os.path.join(folder, new_name)
            os.rename(old_path, new_path)
            print(f"  {filename} → {new_name}")

    format_numbers(demo_dir, "img_", 3)

    print("\n【重命名后】")
    for f in sorted(os.listdir(demo_dir)):
        print(f"  📄 {f}")

    # 清理
    shutil.rmtree(demo_dir)
    print("\n🧹 已清理演示文件")

    print("\n【代码模板】")
    print("-" * 40)
    print("""
# 批量重命名通用模板
import os

folder = "你的文件夹路径"

for filename in os.listdir(folder):
    old_path = os.path.join(folder, filename)

    # 跳过文件夹
    if not os.path.isfile(old_path):
        continue

    # 分离文件名和扩展名
    name, ext = os.path.splitext(filename)

    # 构建新文件名（根据需要修改这里）
    new_name = "前缀_" + name + "_后缀" + ext
    new_path = os.path.join(folder, new_name)

    # 执行重命名
    os.rename(old_path, new_path)
    print(f"重命名：{filename} → {new_name}")
    """)

    print("=" * 60)
    print()


# ============================================================
#                  第4节：文件整理器
# ============================================================

def section_4_file_organizer():
    """
    第4节：文件整理器

    【学习目标】
    创建一个能自动按文件类型分类整理的工具

    【应用场景】
    1. 整理下载文件夹
    2. 整理桌面
    3. 整理照片文件夹

    【核心知识点】
    1. 字典的使用
    2. 文件扩展名判断
    3. 文件夹创建和移动
    """
    print("=" * 60)
    print("第4节：文件整理器")
    print("=" * 60)

    # 创建演示文件夹
    demo_dir = "test_files/messy_folder"
    os.makedirs(demo_dir, exist_ok=True)

    print("\n【场景】假设你的下载文件夹很乱...")
    print("-" * 40)

    # 创建各种类型的测试文件
    test_files = {
        "图片": ["photo1.jpg", "wallpaper.png", "icon.gif"],
        "文档": ["report.docx", "notes.txt", "resume.pdf"],
        "视频": ["movie.mp4", "clip.avi"],
        "音乐": ["song.mp3", "music.wav"],
        "压缩包": ["backup.zip", "files.rar"]
    }

    for category, files in test_files.items():
        for f in files:
            with open(os.path.join(demo_dir, f), "w") as file:
                file.write(f"This is a {category} file")

    print("整理前的文件：")
    for f in sorted(os.listdir(demo_dir)):
        print(f"  📄 {f}")

    print("\n【开始整理】")
    print("-" * 40)

    def organize_files(folder):
        """
        按文件类型整理文件

        参数：
            folder: 要整理的文件夹路径

        返回：
            整理后的分类统计

        示例调用：
            result = organize_files("Downloads")
            print(result)
        """
        # 定义文件类型分类规则
        # key: 文件夹名称
        # value: 该类型包含的扩展名列表
        categories = {
            "图片": [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg"],
            "文档": [".txt", ".doc", ".docx", ".pdf", ".ppt", ".pptx", ".xls", ".xlsx"],
            "视频": [".mp4", ".avi", ".mov", ".mkv", ".flv"],
            "音乐": [".mp3", ".wav", ".flac", ".aac", ".ogg"],
            "压缩包": [".zip", ".rar", ".7z", ".tar", ".gz"],
            "代码": [".py", ".js", ".html", ".css", ".java", ".cpp"],
            "其他": []  # 其他类型放在这里
        }

        # 统计每种类型的数量
        stats = {cat: 0 for cat in categories}

        # 遍历文件夹中的所有文件
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)

            # 跳过文件夹
            if not os.path.isfile(file_path):
                continue

            # 获取文件扩展名（小写）
            ext = os.path.splitext(filename)[1].lower()

            # 确定文件类型
            target_folder = "其他"
            for category, extensions in categories.items():
                if ext in extensions:
                    target_folder = category
                    break

            # 创建目标文件夹
            target_path = os.path.join(folder, target_folder)
            os.makedirs(target_path, exist_ok=True)

            # 移动文件
            new_path = os.path.join(target_path, filename)
            shutil.move(file_path, new_path)

            # 更新统计
            stats[target_folder] += 1
            print(f"  📁 {filename} → {target_folder}/")

        return stats

    # 执行整理
    result = organize_files(demo_dir)

    print("\n【整理结果】")
    print("-" * 40)

    print("整理后的文件夹结构：")
    for item in sorted(os.listdir(demo_dir)):
        item_path = os.path.join(demo_dir, item)
        if os.path.isdir(item_path):
            files = os.listdir(item_path)
            print(f"  📂 {item}/ ({len(files)}个文件)")
            for f in files:
                print(f"      📄 {f}")

    print("\n统计：")
    for category, count in result.items():
        if count > 0:
            print(f"  {category}: {count}个文件")

    # 清理
    shutil.rmtree(demo_dir)
    print("\n🧹 已清理演示文件")

    print("\n【核心代码】")
    print("-" * 40)
    print("""
# 文件分类规则（字典）
categories = {
    "图片": [".jpg", ".png", ".gif"],
    "文档": [".txt", ".pdf", ".docx"],
    "视频": [".mp4", ".avi"],
    "音乐": [".mp3", ".wav"],
}

# 判断文件类型
ext = os.path.splitext(filename)[1].lower()
for category, extensions in categories.items():
    if ext in extensions:
        # 找到类型，移动到对应文件夹
        break
    """)

    print("=" * 60)
    print()


# ============================================================
#                  第5节：Excel自动化
# ============================================================

def section_5_excel_automation():
    """
    第5节：Excel自动化

    【学习目标】
    学会用Python读写Excel文件

    【前置要求】
    需要安装openpyxl库：
    pip install openpyxl

    【核心知识点】
    1. 创建和打开Excel文件
    2. 读取单元格数据
    3. 写入单元格数据
    4. 批量处理数据
    """
    print("=" * 60)
    print("第5节：Excel自动化")
    print("=" * 60)

    print("\n【准备工作】")
    print("-" * 40)
    print("本节需要安装 openpyxl 库")
    print("安装命令：pip install openpyxl")

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        print("✅ openpyxl 已安装，继续学习！")

    except ImportError:
        print("❌ 未安装 openpyxl")
        print("请先运行：pip install openpyxl")
        print("安装后重新运行本教程")
        print("=" * 60)
        print()
        return

    # 创建演示目录
    demo_dir = "test_files/excel_demo"
    os.makedirs(demo_dir, exist_ok=True)

    print("\n【示例1】创建Excel文件并写入数据")
    print("-" * 40)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩表"

    # 写入表头
    headers = ["姓名", "语文", "数学", "英语", "总分", "平均分"]
    ws.append(headers)

    # 写入学生数据
    students = [
        ["小明", 85, 92, 88],
        ["小红", 92, 88, 95],
        ["小刚", 78, 85, 82],
        ["小芳", 88, 90, 92],
        ["小华", 90, 95, 88]
    ]

    for student in students:
        # 计算总分和平均分
        total = sum(student[1:])
        avg = total / 3
        student.extend([total, round(avg, 1)])
        ws.append(student)

    # 保存文件
    excel_file = os.path.join(demo_dir, "student_scores.xlsx")
    wb.save(excel_file)
    print(f"✅ 创建文件：{excel_file}")

    print("\n【示例2】读取Excel数据")
    print("-" * 40)

    # 打开文件
    wb = load_workbook(excel_file)
    ws = wb.active

    print("学生成绩数据：")
    print("-" * 50)

    # 读取所有数据
    for row in ws.iter_rows(values_only=True):
        # 格式化输出
        if row[0] == "姓名":  # 表头
            print(f"{row[0]:<8} {row[1]:<6} {row[2]:<6} {row[3]:<6} {row[4]:<6} {row[5]:<6}")
            print("-" * 50)
        else:
            print(f"{row[0]:<8} {row[1]:<6} {row[2]:<6} {row[3]:<6} {row[4]:<6} {row[5]:<6}")

    print("\n【示例3】设置单元格样式")
    print("-" * 40)

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")  # 白色加粗
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色背景
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:  # 第一行（表头）
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 设置所有单元格居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 8

    # 保存修改
    wb.save(excel_file)
    print("✅ 样式设置完成！")

    print("\n【示例4】数据分析")
    print("-" * 40)

    # 统计分析
    wb = load_workbook(excel_file)
    ws = wb.active

    # 获取数据（跳过表头）
    scores = {"语文": [], "数学": [], "英语": []}
    for row in ws.iter_rows(min_row=2, values_only=True):
        scores["语文"].append(row[1])
        scores["数学"].append(row[2])
        scores["英语"].append(row[3])

    print("各科目统计：")
    for subject, score_list in scores.items():
        avg = sum(score_list) / len(score_list)
        max_score = max(score_list)
        min_score = min(score_list)
        print(f"  {subject}: 平均分={avg:.1f}, 最高分={max_score}, 最低分={min_score}")

    # 清理
    shutil.rmtree(demo_dir)
    print("\n🧹 已清理演示文件")

    print("\n【核心代码】")
    print("-" * 40)
    print("""
from openpyxl import Workbook, load_workbook

# 创建新Excel
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# 写入数据
ws["A1"] = "标题"           # 单个单元格
ws.append(["A", "B", "C"])   # 追加一行

# 保存
wb.save("文件名.xlsx")

# 读取Excel
wb = load_workbook("文件名.xlsx")
ws = wb.active

# 读取数据
value = ws["A1"].value                     # 单个单元格
for row in ws.iter_rows(values_only=True): # 遍历所有行
    print(row)
    """)

    print("=" * 60)
    print()


# ============================================================
#                  第6节：文本处理
# ============================================================

def section_6_text_processing():
    """
    第6节：文本处理

    【学习目标】
    学会用Python处理文本文件，进行批量替换和分析

    【应用场景】
    1. 批量替换文档中的文字
    2. 日志文件分析
    3. 数据提取和清洗
    """
    print("=" * 60)
    print("第6节：文本处理")
    print("=" * 60)

    # 创建演示目录
    demo_dir = "test_files/text_demo"
    os.makedirs(demo_dir, exist_ok=True)

    print("\n【示例1】批量文本替换")
    print("-" * 40)

    # 创建测试文件
    test_file = os.path.join(demo_dir, "article.txt")
    with open(test_file, "w", encoding="utf-8") as f:
        f.write("""Python是一门很棒的编程语言。
学习Python可以让你变得更强大。
Python可以用来做网站、游戏、数据分析。
大家都应该学习Python！
        """)

    def replace_in_file(file_path, old_text, new_text):
        """
        在文件中批量替换文本

        参数：
            file_path: 文件路径
            old_text: 要替换的文本
            new_text: 替换成的新文本

        返回：
            替换的次数

        示例调用：
            count = replace_in_file("test.txt", "Python", "Java")
        """
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        count = content.count(old_text)
        new_content = content.replace(old_text, new_text)

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(new_content)

        return count

    print("原始内容：")
    with open(test_file, "r", encoding="utf-8") as f:
        print(f.read())

    count = replace_in_file(test_file, "Python", "🐍 Python")
    print(f"\n替换了 {count} 处 'Python' → '🐍 Python'")
    print("\n替换后内容：")
    with open(test_file, "r", encoding="utf-8") as f:
        print(f.read())

    print("\n【示例2】日志文件分析")
    print("-" * 40)

    # 创建模拟日志文件
    log_file = os.path.join(demo_dir, "server.log")
    with open(log_file, "w", encoding="utf-8") as f:
        f.write("""[2024-01-15 08:00:01] INFO: 服务器启动
[2024-01-15 08:00:02] INFO: 连接数据库成功
[2024-01-15 08:01:15] ERROR: 用户登录失败 - 密码错误
[2024-01-15 08:02:30] INFO: 用户张三登录成功
[2024-01-15 08:05:00] ERROR: 文件上传失败 - 磁盘空间不足
[2024-01-15 08:10:00] WARNING: 内存使用率85%
[2024-01-15 08:15:00] INFO: 执行定时任务
[2024-01-15 08:20:00] ERROR: 数据库连接超时
[2024-01-15 08:25:00] INFO: 用户李四登录成功
[2024-01-15 08:30:00] WARNING: CPU使用率90%
        """)

    def analyze_log(file_path):
        """
        分析日志文件，统计各类日志数量

        参数：
            file_path: 日志文件路径

        返回：
            各级别日志的数量统计

        示例调用：
            stats = analyze_log("server.log")
        """
        stats = {"INFO": 0, "WARNING": 0, "ERROR": 0}
        errors = []

        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                if "INFO" in line:
                    stats["INFO"] += 1
                elif "WARNING" in line:
                    stats["WARNING"] += 1
                elif "ERROR" in line:
                    stats["ERROR"] += 1
                    errors.append(line.strip())

        return stats, errors

    stats, errors = analyze_log(log_file)

    print("日志统计结果：")
    print(f"  ✅ INFO: {stats['INFO']}条")
    print(f"  ⚠️  WARNING: {stats['WARNING']}条")
    print(f"  ❌ ERROR: {stats['ERROR']}条")

    print("\n错误详情：")
    for error in errors:
        print(f"  {error}")

    print("\n【示例3】提取数据")
    print("-" * 40)

    # 创建包含数据的文件
    data_file = os.path.join(demo_dir, "data.txt")
    with open(data_file, "w", encoding="utf-8") as f:
        f.write("""商品: 苹果, 价格: 5.5元
商品: 香蕉, 价格: 3.2元
商品: 橙子, 价格: 4.8元
商品: 葡萄, 价格: 12.5元
商品: 西瓜, 价格: 8.0元
        """)

    def extract_prices(file_path):
        """
        从文件中提取价格数据

        参数：
            file_path: 文件路径

        返回：
            商品名称和价格的列表

        示例调用：
            items = extract_prices("data.txt")
        """
        items = []

        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                if "商品:" in line and "价格:" in line:
                    # 简单的字符串分割提取
                    parts = line.strip().split(", ")
                    name = parts[0].replace("商品: ", "")
                    price = parts[1].replace("价格: ", "").replace("元", "")
                    items.append((name, float(price)))

        return items

    items = extract_prices(data_file)

    print("提取的商品数据：")
    total = 0
    for name, price in items:
        print(f"  {name}: ¥{price}")
        total += price
    print(f"\n总价: ¥{total}")
    print(f"平均价格: ¥{total/len(items):.2f}")

    # 清理
    shutil.rmtree(demo_dir)
    print("\n🧹 已清理演示文件")

    print("\n【核心代码】")
    print("-" * 40)
    print("""
# 读取文件
with open("文件.txt", "r", encoding="utf-8") as f:
    content = f.read()

# 替换文本
new_content = content.replace("旧文本", "新文本")

# 写入文件
with open("文件.txt", "w", encoding="utf-8") as f:
    f.write(new_content)

# 按行读取
with open("文件.txt", "r", encoding="utf-8") as f:
    for line in f:
        if "关键词" in line:
            print(line.strip())
    """)

    print("=" * 60)
    print()


# ============================================================
#                  第7节：定时任务
# ============================================================

def section_7_scheduled_tasks():
    """
    第7节：定时任务

    【学习目标】
    学会让程序在指定时间自动执行

    【应用场景】
    1. 定时提醒喝水、休息
    2. 定时备份文件
    3. 定时发送通知

    【核心知识点】
    1. schedule库的使用
    2. time模块的sleep函数
    3. while循环
    """
    print("=" * 60)
    print("第7节：定时任务")
    print("=" * 60)

    print("\n【准备工作】")
    print("-" * 40)
    print("本节需要安装 schedule 库")
    print("安装命令：pip install schedule")

    try:
        import schedule
        print("✅ schedule 已安装，继续学习！")
    except ImportError:
        print("❌ 未安装 schedule")
        print("请先运行：pip install schedule")
        print("\n以下是代码示例（需要安装后才能运行）：")


    print("\n【核心概念】")
    print("-" * 40)
    print("""
定时任务 = 让程序按计划自动执行

常见的定时方式：
  - 每隔N秒/分钟/小时执行
  - 每天固定时间执行
  - 每周固定时间执行
  - 特定日期执行
    """)

    print("\n【示例1】简单定时器")
    print("-" * 40)

    def simple_timer(seconds, message):
        """
        简单的倒计时定时器

        参数：
            seconds: 倒计时秒数
            message: 时间到显示的消息

        示例调用：
            simple_timer(5, "时间到！该休息了！")
        """
        print(f"⏰ 设置 {seconds} 秒后的提醒...")
        for i in range(seconds, 0, -1):
            print(f"  倒计时: {i}秒", end="\r")
            time.sleep(1)
        print(f"\n🔔 {message}")

    # 演示（只等3秒）
    print("演示：3秒后提醒")
    simple_timer(3, "该喝水了！💧")

    print("\n【示例2】使用schedule库（代码示例）")
    print("-" * 40)

    schedule_code = '''
import schedule
import time

# 定义任务函数
def drink_water():
    print("💧 该喝水了！")

def take_break():
    print("🏃 该休息一下了！")

def backup_files():
    print("💾 正在备份文件...")

# 设置定时任务
schedule.every(1).hours.do(drink_water)      # 每1小时
schedule.every(2).hours.do(take_break)        # 每2小时
schedule.every().day.at("18:00").do(backup)   # 每天18:00

# 运行定时任务
print("定时任务已启动，按Ctrl+C停止")
while True:
    schedule.run_pending()  # 检查是否有任务需要执行
    time.sleep(1)           # 等待1秒
'''
    print(schedule_code)

    print("\n【常用定时设置】")
    print("-" * 40)
    print("""
schedule.every(10).seconds.do(任务)       # 每10秒
schedule.every(10).minutes.do(任务)       # 每10分钟
schedule.every().hour.do(任务)            # 每小时
schedule.every().day.at("10:30").do(任务) # 每天10:30
schedule.every().monday.do(任务)          # 每周一
schedule.every().wednesday.at("14:00").do(任务) # 每周三14:00
    """)

    print("\n【实际应用：健康提醒助手】")
    print("-" * 40)

    health_reminder_code = '''
import schedule
import time
from datetime import datetime

def remind_water():
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 💧 该喝水了！保持水分很重要")

def remind_eye():
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 👀 让眼睛休息一下，看看远处")

def remind_stretch():
    print(f"[{datetime.now().strftime('%H:%M:%S')]) 🧘 站起来活动活动")

# 设置提醒（演示用较短间隔）
schedule.every(30).seconds.do(remind_water)
schedule.every(45).seconds.do(remind_eye)
schedule.every(60).seconds.do(remind_stretch)

print("🏥 健康提醒助手已启动！")
print("提醒间隔：")
print("  - 喝水：每30秒")
print("  - 护眼：每45秒")
print("  - 活动：每60秒")
print("按Ctrl+C停止\\n")

try:
    while True:
        schedule.run_pending()
        time.sleep(1)
except KeyboardInterrupt:
    print("\\n👋 提醒助手已停止")
'''
    print(health_reminder_code)

    print("=" * 60)
    print()


# ============================================================
#                  第8节：综合案例
# ============================================================

def section_8_comprehensive_case():
    """
    第8节：综合案例

    【学习目标】
    综合运用前面学到的知识，创建一个实用的工具

    【案例】智能作业文件整理器
    功能：
    1. 自动按科目分类文件
    2. 按日期重命名
    3. 生成整理报告
    """
    print("=" * 60)
    print("第8节：综合案例 - 智能作业整理器")
    print("=" * 60)

    print("\n【案例介绍】")
    print("-" * 40)
    print("""
假设你是学生，每周都会收到各科目的作业文件：
  - math_homework_001.pdf
  - chinese_essay.docx
  - english_reading.txt
  ...

这些文件散落在下载文件夹里，很乱！
让我们创建一个工具自动整理这些文件。
    """)

    # 创建演示环境
    demo_dir = "test_files/homework_demo"
    os.makedirs(demo_dir, exist_ok=True)

    # 创建模拟作业文件
    homework_files = [
        "math_homework_001.pdf",
        "math_quiz.pdf",
        "chinese_essay.docx",
        "chinese_notes.txt",
        "english_reading.pdf",
        "english_vocab.xlsx",
        "physics_lab.pdf",
        "chemistry_report.docx",
        "history_timeline.pptx"
    ]

    for f in homework_files:
        with open(os.path.join(demo_dir, f), "w") as file:
            file.write(f"这是 {f} 的内容")

    print("整理前的文件：")
    for f in sorted(os.listdir(demo_dir)):
        print(f"  📄 {f}")

    print("\n【开始整理】")
    print("-" * 40)

    def organize_homework(folder):
        """
        智能整理作业文件

        功能：
        1. 按科目分类
        2. 添加日期前缀
        3. 生成整理报告

        参数：
            folder: 要整理的文件夹路径

        返回：
            整理统计信息
        """
        # 科目关键词映射
        subjects = {
            "数学": ["math"],
            "语文": ["chinese", "chinese_essay", "作文"],
            "英语": ["english"],
            "物理": ["physics"],
            "化学": ["chemistry"],
            "历史": ["history"],
            "地理": ["geography"],
            "生物": ["biology"],
            "政治": ["politics"]
        }

        # 获取今天的日期
        today = datetime.now().strftime("%Y%m%d")

        # 统计
        stats = {subject: 0 for subject in subjects}
        stats["其他"] = 0

        # 处理每个文件
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)

            if not os.path.isfile(file_path):
                continue

            # 判断科目
            subject = "其他"
            filename_lower = filename.lower()

            for subj, keywords in subjects.items():
                if any(keyword in filename_lower for keyword in keywords):
                    subject = subj
                    break

            # 创建科目文件夹
            subject_folder = os.path.join(folder, subject)
            os.makedirs(subject_folder, exist_ok=True)

            # 新文件名：日期_原文件名
            new_name = f"{today}_{filename}"
            new_path = os.path.join(subject_folder, new_name)

            # 移动文件
            shutil.move(file_path, new_path)

            # 更新统计
            stats[subject] += 1
            print(f"  📁 {filename} → {subject}/{new_name}")

        return stats

    # 执行整理
    result = organize_homework(demo_dir)

    print("\n【整理结果】")
    print("-" * 40)

    print("文件夹结构：")
    for item in sorted(os.listdir(demo_dir)):
        item_path = os.path.join(demo_dir, item)
        if os.path.isdir(item_path):
            files = os.listdir(item_path)
            print(f"\n  📂 {item}/ ({len(files)}个文件)")
            for f in files:
                print(f"      📄 {f}")

    print("\n\n整理统计：")
    print("-" * 30)
    total = 0
    for subject, count in result.items():
        if count > 0:
            print(f"  {subject}: {count}个文件")
            total += 1
    print(f"\n  共整理 {sum(result.values())} 个文件，分为 {total} 个科目")

    # 生成报告
    report_file = os.path.join(demo_dir, "整理报告.txt")
    with open(report_file, "w", encoding="utf-8") as f:
        f.write("=" * 40 + "\n")
        f.write("作业文件整理报告\n")
        f.write("=" * 40 + "\n\n")
        f.write(f"整理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("统计信息:\n")
        f.write("-" * 30 + "\n")
        for subject, count in result.items():
            if count > 0:
                f.write(f"  {subject}: {count}个文件\n")
        f.write(f"\n总计: {sum(result.values())}个文件\n")

    print(f"\n📝 已生成整理报告: {report_file}")

    # 清理
    shutil.rmtree(demo_dir)
    print("\n🧹 已清理演示文件")

    print("\n【学到的知识点】")
    print("-" * 40)
    print("""
1. 文件操作：创建文件夹、移动文件
2. 字符串处理：判断文件名中的关键词
3. 日期处理：获取当前日期并格式化
4. 字典使用：统计分类信息
5. 文件读写：生成报告文件

这个案例把前7节的内容都用上了！
    """)

    print("\n【完整代码】请查看 scripts/file_organizer.py")
    print("=" * 60)
    print()


# ============================================================
#                      主程序
# ============================================================

def print_menu():
    """打印菜单"""
    print("""
╔════════════════════════════════════════════════════════════╗
║           自动化脚本快速入门教程 v1.0                        ║
║                  适合初中生学习                              ║
╠════════════════════════════════════════════════════════════╣
║  0. 运行全部教程                                            ║
║  1. 什么是自动化？（概念介绍）                               ║
║  2. 文件操作自动化                                          ║
║  3. 批量重命名                                              ║
║  4. 文件整理器                                              ║
║  5. Excel自动化                                             ║
║  6. 文本处理                                                ║
║  7. 定时任务                                                ║
║  8. 综合案例                                                ║
║  q. 退出                                                    ║
╚════════════════════════════════════════════════════════════╝
    """)


def main():
    """主函数"""
    sections = {
        "1": section_1_what_is_automation,
        "2": section_2_file_operations,
        "3": section_3_batch_rename,
        "4": section_4_file_organizer,
        "5": section_5_excel_automation,
        "6": section_6_text_processing,
        "7": section_7_scheduled_tasks,
        "8": section_8_comprehensive_case,
    }

    while True:
        print_menu()
        choice = input("请选择要学习的章节 (0-8, q退出): ").strip()

        if choice == "q":
            print("\n👋 感谢学习！继续加油！")
            break
        elif choice == "0":
            # 运行全部教程
            print("\n🚀 开始完整教程...\n")
            for i in range(1, 9):
                sections[str(i)]()
                input("\n按回车键继续下一节...")
        elif choice in sections:
            print()
            sections[choice]()
            input("\n按回车键返回菜单...")
        else:
            print("❌ 无效选择，请重新输入")


if __name__ == "__main__":
    main()
