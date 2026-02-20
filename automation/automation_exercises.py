#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动化脚本练习题
================
配合 automation_quick_start.py 教程使用

每个练习都包含：
  - 题目描述
  - 提示
  - 答案（运行后可查看）

运行方式：
    python automation_exercises.py
"""

import os
import shutil
import random
import string
from datetime import datetime

# ============================================================
#                    练习环境设置
# ============================================================

def setup_exercise_env():
    """设置练习环境"""
    exercise_dir = "test_files/exercises"
    if os.path.exists(exercise_dir):
        shutil.rmtree(exercise_dir)
    os.makedirs(exercise_dir, exist_ok=True)
    return exercise_dir


# ============================================================
#                    第2节练习：文件操作
# ============================================================

def exercise_2_1():
    """
    练习2.1：创建文件夹

    题目：在 test_files/exercises 目录下创建一个名为 "my_folder" 的文件夹

    提示：使用 os.makedirs() 函数
    """
    print("=" * 60)
    print("练习2.1：创建文件夹")
    print("=" * 60)
    print("\n题目：在 test_files/exercises 目录下创建名为 'my_folder' 的文件夹")
    print("\n提示：使用 os.makedirs() 函数")
    print("\n请尝试写出代码：")
    print("-" * 40)

    # 参考答案
    answer = '''
# 参考答案
import os

folder_path = "test_files/exercises/my_folder"
os.makedirs(folder_path, exist_ok=True)
print(f"已创建文件夹：{folder_path}")

# 验证
if os.path.exists(folder_path):
    print("✅ 文件夹创建成功！")
'''

    input("按回车查看答案...")
    print(answer)

    # 执行演示
    print("\n运行结果：")
    folder_path = "test_files/exercises/my_folder"
    os.makedirs(folder_path, exist_ok=True)
    print(f"已创建文件夹：{folder_path}")
    if os.path.exists(folder_path):
        print("✅ 文件夹创建成功！")


def exercise_2_2():
    """
    练习2.2：创建文件并写入内容

    题目：创建一个名为 "diary.txt" 的文件，写入今天的日期和"今天学习了Python自动化！"

    提示：使用 open() 函数和 with 语句
    """
    print("=" * 60)
    print("练习2.2：创建文件并写入内容")
    print("=" * 60)
    print("\n题目：创建 diary.txt，写入日期和学习记录")
    print("\n提示：使用 with open(...) as f: 和 f.write()")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
from datetime import datetime

file_path = "test_files/exercises/diary.txt"
today = datetime.now().strftime("%Y-%m-%d")

with open(file_path, "w", encoding="utf-8") as f:
    f.write(f"日期：{today}\\n")
    f.write("今天学习了Python自动化！\\n")

print(f"已创建文件：{file_path}")

# 验证：读取并显示内容
with open(file_path, "r", encoding="utf-8") as f:
    print("文件内容：")
    print(f.read())
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    file_path = "test_files/exercises/diary.txt"
    today = datetime.now().strftime("%Y-%m-%d")
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(f"日期：{today}\n")
        f.write("今天学习了Python自动化！\n")
    print(f"已创建文件：{file_path}")
    with open(file_path, "r", encoding="utf-8") as f:
        print("文件内容：")
        print(f.read())


def exercise_2_3():
    """
    练习2.3：复制文件

    题目：将 diary.txt 复制一份，命名为 diary_backup.txt

    提示：使用 shutil.copy()
    """
    print("=" * 60)
    print("练习2.3：复制文件")
    print("=" * 60)
    print("\n题目：将 diary.txt 复制为 diary_backup.txt")
    print("\n提示：使用 shutil.copy(源文件, 目标文件)")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import shutil

source = "test_files/exercises/diary.txt"
target = "test_files/exercises/diary_backup.txt"

shutil.copy(source, target)
print(f"已复制：{source} → {target}")

# 验证
import os
if os.path.exists(target):
    print("✅ 复制成功！")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    source = "test_files/exercises/diary.txt"
    target = "test_files/exercises/diary_backup.txt"
    shutil.copy(source, target)
    print(f"已复制：{source} → {target}")
    if os.path.exists(target):
        print("✅ 复制成功！")


# ============================================================
#                    第3节练习：批量重命名
# ============================================================

def exercise_3_1():
    """
    练习3.1：添加前缀

    题目：给文件夹中的所有 .txt 文件添加 "note_" 前缀
    """
    print("=" * 60)
    print("练习3.1：给文件添加前缀")
    print("=" * 60)

    # 创建测试文件
    test_dir = "test_files/exercises/rename_test"
    os.makedirs(test_dir, exist_ok=True)
    for i in range(1, 4):
        with open(os.path.join(test_dir, f"file{i}.txt"), "w") as f:
            f.write(f"内容{i}")
        with open(os.path.join(test_dir, f"data{i}.csv"), "w") as f:
            f.write(f"data,{i}")

    print("\n原始文件：")
    for f in os.listdir(test_dir):
        print(f"  📄 {f}")

    print("\n题目：给所有 .txt 文件添加 'note_' 前缀")
    print("\n提示：使用 os.listdir(), endswith(), os.rename()")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import os

folder = "test_files/exercises/rename_test"

for filename in os.listdir(folder):
    if filename.endswith(".txt"):
        old_path = os.path.join(folder, filename)
        new_name = "note_" + filename
        new_path = os.path.join(folder, new_name)
        os.rename(old_path, new_path)
        print(f"重命名：{filename} → {new_name}")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    folder = "test_files/exercises/rename_test"
    for filename in os.listdir(folder):
        if filename.endswith(".txt"):
            old_path = os.path.join(folder, filename)
            new_name = "note_" + filename
            new_path = os.path.join(folder, new_name)
            os.rename(old_path, new_path)
            print(f"重命名：{filename} → {new_name}")

    print("\n重命名后：")
    for f in sorted(os.listdir(folder)):
        print(f"  📄 {f}")


def exercise_3_2():
    """
    练习3.2：修改扩展名

    题目：将所有 .csv 文件改为 .data 文件
    """
    print("=" * 60)
    print("练习3.2：修改文件扩展名")
    print("=" * 60)

    # 恢复测试文件
    test_dir = "test_files/exercises/rename_test"
    for f in os.listdir(test_dir):
        if f.endswith(".csv"):
            os.remove(os.path.join(test_dir, f))
    for i in range(1, 4):
        with open(os.path.join(test_dir, f"data{i}.csv"), "w") as f:
            f.write("data")

    print("\n原始文件：")
    csv_files = [f for f in os.listdir(test_dir) if f.endswith(".csv")]
    for f in csv_files:
        print(f"  📄 {f}")

    print("\n题目：将所有 .csv 文件改为 .data 文件")
    print("\n提示：使用 os.path.splitext() 分离文件名和扩展名")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import os

folder = "test_files/exercises/rename_test"

for filename in os.listdir(folder):
    if filename.endswith(".csv"):
        old_path = os.path.join(folder, filename)
        name = os.path.splitext(filename)[0]  # 获取不带扩展名的文件名
        new_name = name + ".data"
        new_path = os.path.join(folder, new_name)
        os.rename(old_path, new_path)
        print(f"重命名：{filename} → {new_name}")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    folder = "test_files/exercises/rename_test"
    for filename in os.listdir(folder):
        if filename.endswith(".csv"):
            old_path = os.path.join(folder, filename)
            name = os.path.splitext(filename)[0]
            new_name = name + ".data"
            new_path = os.path.join(folder, new_name)
            os.rename(old_path, new_path)
            print(f"重命名：{filename} → {new_name}")

    print("\n修改后：")
    for f in sorted(os.listdir(folder)):
        if f.endswith(".data"):
            print(f"  📄 {f}")


# ============================================================
#                    第4节练习：文件整理
# ============================================================

def exercise_4_1():
    """
    练习4.1：分类整理文件

    题目：将文件按扩展名分类到不同文件夹
    """
    print("=" * 60)
    print("练习4.1：分类整理文件")
    print("=" * 60)

    # 创建测试文件
    test_dir = "test_files/exercises/organize_test"
    if os.path.exists(test_dir):
        shutil.rmtree(test_dir)
    os.makedirs(test_dir)

    # 创建各种类型文件
    files = ["photo1.jpg", "photo2.png", "doc1.txt", "doc2.pdf", "song.mp3"]
    for f in files:
        with open(os.path.join(test_dir, f), "w") as file:
            file.write("test")

    print("\n原始文件：")
    for f in os.listdir(test_dir):
        print(f"  📄 {f}")

    print("\n题目：将文件按类型分类")
    print("  - 图片(.jpg, .png) → images 文件夹")
    print("  - 文档(.txt, .pdf) → documents 文件夹")
    print("  - 其他 → others 文件夹")
    print("\n提示：使用字典定义分类规则，然后遍历文件并移动")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import os
import shutil

folder = "test_files/exercises/organize_test"

# 定义分类规则
categories = {
    "images": [".jpg", ".png", ".gif"],
    "documents": [".txt", ".pdf", ".docx"],
    "others": []
}

for filename in os.listdir(folder):
    file_path = os.path.join(folder, filename)

    if not os.path.isfile(file_path):
        continue

    # 获取扩展名
    ext = os.path.splitext(filename)[1].lower()

    # 确定目标文件夹
    target = "others"
    for category, extensions in categories.items():
        if ext in extensions:
            target = category
            break

    # 创建目标文件夹并移动
    target_folder = os.path.join(folder, target)
    os.makedirs(target_folder, exist_ok=True)
    shutil.move(file_path, os.path.join(target_folder, filename))
    print(f"  {filename} → {target}/")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    folder = "test_files/exercises/organize_test"
    categories = {
        "images": [".jpg", ".png", ".gif"],
        "documents": [".txt", ".pdf", ".docx"],
        "others": []
    }

    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        if not os.path.isfile(file_path):
            continue
        ext = os.path.splitext(filename)[1].lower()
        target = "others"
        for category, extensions in categories.items():
            if ext in extensions:
                target = category
                break
        target_folder = os.path.join(folder, target)
        os.makedirs(target_folder, exist_ok=True)
        shutil.move(file_path, os.path.join(target_folder, filename))
        print(f"  {filename} → {target}/")

    print("\n整理后结构：")
    for item in sorted(os.listdir(folder)):
        item_path = os.path.join(folder, item)
        if os.path.isdir(item_path):
            print(f"  📂 {item}/")
            for f in os.listdir(item_path):
                print(f"      📄 {f}")


# ============================================================
#                    第5节练习：Excel处理
# ============================================================

def exercise_5_1():
    """
    练习5.1：创建成绩表

    题目：创建一个Excel文件，包含3个学生的成绩
    """
    print("=" * 60)
    print("练习5.1：创建Excel成绩表")
    print("=" * 60)

    print("\n题目：创建一个Excel文件，包含以下数据：")
    print("  姓名  | 语文 | 数学 | 英语")
    print("  ------|------|------|------")
    print("  小明  |  85  |  92  |  88")
    print("  小红  |  90  |  88  |  95")
    print("  小刚  |  78  |  85  |  82")

    print("\n提示：需要先 pip install openpyxl")
    print("     使用 Workbook() 创建工作簿，ws.append() 添加数据")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
from openpyxl import Workbook

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "成绩表"

# 写入表头
headers = ["姓名", "语文", "数学", "英语"]
ws.append(headers)

# 写入数据
students = [
    ["小明", 85, 92, 88],
    ["小红", 90, 88, 95],
    ["小刚", 78, 85, 82]
]

for student in students:
    ws.append(student)

# 保存文件
wb.save("test_files/exercises/成绩表.xlsx")
print("✅ 已创建 test_files/exercises/成绩表.xlsx")
'''

    input("按回车查看答案...")
    print(answer)

    try:
        from openpyxl import Workbook

        print("\n运行结果：")
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩表"
        ws.append(["姓名", "语文", "数学", "英语"])
        students = [["小明", 85, 92, 88], ["小红", 90, 88, 95], ["小刚", 78, 85, 82]]
        for student in students:
            ws.append(student)
        wb.save("test_files/exercises/成绩表.xlsx")
        print("✅ 已创建 test_files/exercises/成绩表.xlsx")
    except ImportError:
        print("❌ 需要先安装 openpyxl: pip install openpyxl")


def exercise_5_2():
    """
    练习5.2：读取Excel并计算平均分

    题目：读取成绩表，计算每个学生的平均分
    """
    print("=" * 60)
    print("练习5.2：读取Excel并计算平均分")
    print("=" * 60)

    print("\n题目：读取上一步创建的成绩表，计算每个学生的平均分")
    print("\n提示：使用 load_workbook() 打开文件，iter_rows() 遍历数据")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
from openpyxl import load_workbook

# 打开文件
wb = load_workbook("test_files/exercises/成绩表.xlsx")
ws = wb.active

print("学生平均分：")
print("-" * 30)

# 遍历数据（跳过表头）
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    scores = row[1:4]  # 语文、数学、英语
    avg = sum(scores) / len(scores)
    print(f"{name}: {avg:.1f}分")
'''

    input("按回车查看答案...")
    print(answer)

    try:
        from openpyxl import load_workbook

        print("\n运行结果：")
        wb = load_workbook("test_files/exercises/成绩表.xlsx")
        ws = wb.active
        print("学生平均分：")
        print("-" * 30)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            scores = row[1:4]
            avg = sum(scores) / len(scores)
            print(f"{name}: {avg:.1f}分")
    except Exception as e:
        print(f"❌ 错误: {e}")


# ============================================================
#                    第6节练习：文本处理
# ============================================================

def exercise_6_1():
    """
    练习6.1：统计单词出现次数

    题目：统计文本中每个单词出现的次数
    """
    print("=" * 60)
    print("练习6.1：统计单词出现次数")
    print("=" * 60)

    # 创建测试文件
    test_file = "test_files/exercises/article.txt"
    with open(test_file, "w", encoding="utf-8") as f:
        f.write("Python是一门很棒的编程语言。Python可以用来做很多事。\n")
        f.write("学习Python很有趣。Python让自动化变得简单。\n")

    print("\n文本内容：")
    with open(test_file, "r", encoding="utf-8") as f:
        print(f.read())

    print("题目：统计 'Python' 出现了几次")
    print("\n提示：使用 count() 方法")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
file_path = "test_files/exercises/article.txt"

with open(file_path, "r", encoding="utf-8") as f:
    content = f.read()

count = content.count("Python")
print(f"'Python' 出现了 {count} 次")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    with open(test_file, "r", encoding="utf-8") as f:
        content = f.read()
    count = content.count("Python")
    print(f"'Python' 出现了 {count} 次")


def exercise_6_2():
    """
    练习6.2：批量替换

    题目：将文本中的"Python"替换为"Python3"
    """
    print("=" * 60)
    print("练习6.2：批量文本替换")
    print("=" * 60)

    print("\n题目：将 article.txt 中的 'Python' 替换为 'Python3'")
    print("\n提示：使用 replace() 方法")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
file_path = "test_files/exercises/article.txt"

# 读取
with open(file_path, "r", encoding="utf-8") as f:
    content = f.read()

# 替换
new_content = content.replace("Python", "Python3")

# 写回
with open(file_path, "w", encoding="utf-8") as f:
    f.write(new_content)

print("替换完成！")
print("\\n替换后内容：")
with open(file_path, "r", encoding="utf-8") as f:
    print(f.read())
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果：")
    file_path = "test_files/exercises/article.txt"
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()
    new_content = content.replace("Python", "Python3")
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(new_content)
    print("替换完成！")
    print("\n替换后内容：")
    with open(file_path, "r", encoding="utf-8") as f:
        print(f.read())


# ============================================================
#                    第7节练习：定时任务
# ============================================================

def exercise_7_1():
    """
    练习7.1：倒计时

    题目：创建一个10秒倒计时
    """
    print("=" * 60)
    print("练习7.1：倒计时器")
    print("=" * 60)

    print("\n题目：创建一个10秒倒计时，每秒显示剩余秒数")
    print("\n提示：使用 time.sleep(1) 和 for 循环")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import time

print("倒计时开始！")
for i in range(10, 0, -1):
    print(f"  剩余 {i} 秒...")
    time.sleep(1)
print("⏰ 时间到！")
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行结果（只演示3秒）：")
    import time
    print("倒计时开始！")
    for i in range(3, 0, -1):
        print(f"  剩余 {i} 秒...")
        time.sleep(1)
    print("⏰ 时间到！")


def exercise_7_2():
    """
    练习7.2：定时提醒代码
    """
    print("=" * 60)
    print("练习7.2：定时提醒（代码设计）")
    print("=" * 60)

    print("\n题目：使用 schedule 库设计一个每5秒提醒一次的程序")
    print("\n提示：使用 schedule.every(5).seconds.do(函数)")
    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import schedule
import time
from datetime import datetime

def remind():
    current_time = datetime.now().strftime("%H:%M:%S")
    print(f"[{current_time}] 🔔 该休息一下了！")

# 设置每5秒执行
schedule.every(5).seconds.do(remind)

print("定时提醒已启动，按Ctrl+C停止")

# 运行循环
try:
    while True:
        schedule.run_pending()
        time.sleep(1)
except KeyboardInterrupt:
    print("\\n已停止")
'''

    input("按回车查看答案...")
    print(answer)
    print("\n💡 注意：这需要安装 schedule 库并完整运行才能看到效果")


# ============================================================
#                    综合练习
# ============================================================

def exercise_comprehensive():
    """
    综合练习：创建一个简单的文件管理器
    """
    print("=" * 60)
    print("综合练习：文件管理器")
    print("=" * 60)

    print("""
【题目】创建一个简单的文件管理器，实现以下功能：
1. 列出指定文件夹的所有文件
2. 按扩展名分类显示
3. 统计每种类型的文件数量

【要求】
- 用户输入文件夹路径
- 显示分类统计结果

【提示】
- 使用 os.listdir() 获取文件列表
- 使用 os.path.splitext() 获取扩展名
- 使用字典统计数量
    """)

    print("\n请尝试写出代码：")
    print("-" * 40)

    answer = '''
# 参考答案
import os
from collections import Counter

def file_manager(folder_path):
    """
    简单文件管理器

    参数：
        folder_path: 文件夹路径
    """
    if not os.path.exists(folder_path):
        print("文件夹不存在！")
        return

    # 获取所有文件
    files = [f for f in os.listdir(folder_path)
             if os.path.isfile(os.path.join(folder_path, f))]

    if not files:
        print("文件夹为空！")
        return

    print(f"\\n文件夹: {folder_path}")
    print("=" * 50)
    print(f"共 {len(files)} 个文件\\n")

    # 统计扩展名
    extensions = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        extensions.append(ext if ext else "(无扩展名)")

    # 计数
    ext_count = Counter(extensions)

    print("文件类型统计：")
    print("-" * 30)
    for ext, count in sorted(ext_count.items(), key=lambda x: -x[1]):
        print(f"  {ext}: {count}个")

    print("\\n文件列表：")
    print("-" * 30)
    for f in sorted(files):
        print(f"  📄 {f}")

# 使用示例
if __name__ == "__main__":
    path = input("请输入文件夹路径: ")
    file_manager(path)
'''

    input("按回车查看答案...")
    print(answer)

    print("\n运行演示（使用当前目录）：")
    import os
    from collections import Counter

    folder = "."
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]

    print(f"\n文件夹: {os.path.abspath(folder)}")
    print("=" * 50)
    print(f"共 {len(files)} 个文件\n")

    extensions = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        extensions.append(ext if ext else "(无扩展名)")

    ext_count = Counter(extensions)

    print("文件类型统计：")
    print("-" * 30)
    for ext, count in sorted(ext_count.items(), key=lambda x: -x[1]):
        print(f"  {ext}: {count}个")


# ============================================================
#                      主菜单
# ============================================================

def print_menu():
    """打印练习菜单"""
    print("""
╔════════════════════════════════════════════════════════════╗
║              自动化脚本练习题                               ║
╠════════════════════════════════════════════════════════════╣
║  第2节：文件操作                                            ║
║    2.1  创建文件夹                                          ║
║    2.2  创建文件并写入内容                                   ║
║    2.3  复制文件                                            ║
║                                                            ║
║  第3节：批量重命名                                          ║
║    3.1  添加前缀                                            ║
║    3.2  修改扩展名                                          ║
║                                                            ║
║  第4节：文件整理                                            ║
║    4.1  分类整理文件                                        ║
║                                                            ║
║  第5节：Excel处理                                           ║
║    5.1  创建成绩表                                          ║
║    5.2  读取并计算平均分                                    ║
║                                                            ║
║  第6节：文本处理                                            ║
║    6.1  统计单词出现次数                                    ║
║    6.2  批量文本替换                                        ║
║                                                            ║
║  第7节：定时任务                                            ║
║    7.1  倒计时器                                            ║
║    7.2  定时提醒代码                                        ║
║                                                            ║
║  综合  文件管理器                                           ║
║  q     退出                                                 ║
╚════════════════════════════════════════════════════════════╝
    """)


def main():
    """主函数"""
    # 设置练习环境
    setup_exercise_env()

    exercises = {
        "2.1": exercise_2_1,
        "2.2": exercise_2_2,
        "2.3": exercise_2_3,
        "3.1": exercise_3_1,
        "3.2": exercise_3_2,
        "4.1": exercise_4_1,
        "5.1": exercise_5_1,
        "5.2": exercise_5_2,
        "6.1": exercise_6_1,
        "6.2": exercise_6_2,
        "7.1": exercise_7_1,
        "7.2": exercise_7_2,
        "综合": exercise_comprehensive,
    }

    while True:
        print_menu()
        choice = input("请选择练习题 (如 2.1, 3.2, 综合, q退出): ").strip()

        if choice == "q":
            print("\n👋 继续练习，加油！")
            break
        elif choice in exercises:
            print()
            exercises[choice]()
            input("\n按回车返回菜单...")
        else:
            print("❌ 无效选择，请重新输入")


if __name__ == "__main__":
    main()
