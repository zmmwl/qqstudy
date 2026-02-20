#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理工具
=============
简单易用的Excel文件处理工具

功能：
- 创建Excel文件
- 读取Excel数据
- 批量处理数据
- 数据统计
- 格式设置

依赖：
    pip install openpyxl

使用方法：
    python excel_processor.py

作者: Python学习小组
"""

import os

# 检查是否安装了openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  需要安装 openpyxl 库")
    print("   安装命令: pip install openpyxl")


def create_workbook(title="Sheet1"):
    """
    创建一个新的Excel工作簿

    参数：
        title: str - 工作表名称，默认"Sheet1"

    返回：
        Workbook对象

    示例调用：
        wb = create_workbook("成绩表")
        wb.save("成绩表.xlsx")
    """
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb


def write_data(workbook, data, sheet_name=None, start_row=1, start_col=1):
    """
    向Excel写入数据

    参数：
        workbook: Workbook - 工作簿对象
        data: list - 二维列表，每个元素是一行数据
        sheet_name: str - 工作表名称（可选）
        start_row: int - 起始行，默认1
        start_col: int - 起始列，默认1

    返回：
        None

    示例调用：
        wb = create_workbook()
        data = [
            ["姓名", "语文", "数学"],
            ["小明", 85, 92],
            ["小红", 90, 88]
        ]
        write_data(wb, data)
        wb.save("成绩表.xlsx")
    """
    if not OPENPYXL_AVAILABLE:
        return

    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)


def read_data(file_path, sheet_name=None, has_header=True):
    """
    从Excel读取数据

    参数：
        file_path: str - Excel文件路径
        sheet_name: str - 工作表名称（可选，默认读取第一个）
        has_header: bool - 是否有表头，默认True

    返回：
        tuple - (headers, data) 如果有表头
        list - data 如果没有表头

    示例调用：
        headers, data = read_data("成绩表.xlsx")
        print(headers)  # ['姓名', '语文', '数学']
        print(data)     # [['小明', 85, 92], ['小红', 90, 88]]
    """
    if not OPENPYXL_AVAILABLE:
        return None

    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        return None

    wb = load_workbook(file_path, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    all_data = []
    for row in ws.iter_rows(values_only=True):
        all_data.append(list(row))

    wb.close()

    if has_header and len(all_data) > 0:
        return all_data[0], all_data[1:]
    return all_data


def set_header_style(workbook, sheet_name=None):
    """
    设置表头样式（加粗、居中、蓝色背景）

    参数：
        workbook: Workbook - 工作簿对象
        sheet_name: str - 工作表名称（可选）

    返回：
        None

    示例调用：
        wb = create_workbook()
        write_data(wb, [["姓名", "成绩"], ["小明", 90]])
        set_header_style(wb)
        wb.save("成绩表.xlsx")
    """
    if not OPENPYXL_AVAILABLE:
        return

    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    # 定义样式
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    # 应用到第一行
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment


def auto_column_width(workbook, sheet_name=None):
    """
    自动调整列宽

    参数：
        workbook: Workbook - 工作簿对象
        sheet_name: str - 工作表名称（可选）

    返回：
        None

    示例调用：
        wb = create_workbook()
        # ... 写入数据 ...
        auto_column_width(wb)
        wb.save("output.xlsx")
    """
    if not OPENPYXL_AVAILABLE:
        return

    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # 最大宽度50
        ws.column_dimensions[column].width = adjusted_width


def add_formula(file_path, column, formula, sheet_name=None):
    """
    在指定列添加公式

    参数：
        file_path: str - Excel文件路径
        column: int - 列号（从1开始）
        formula: str - 公式字符串（如 "=SUM(B2:D2)"）
        sheet_name: str - 工作表名称（可选）

    返回：
        None

    示例调用：
        # 在E列添加总分公式
        add_formula("成绩表.xlsx", 5, "=SUM(B2:D2)")
    """
    if not OPENPYXL_AVAILABLE:
        return

    wb = load_workbook(file_path)

    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = wb.active

    # 获取数据行数
    max_row = ws.max_row

    # 从第二行开始添加公式
    for row in range(2, max_row + 1):
        cell_formula = formula.replace("2", str(row))
        ws.cell(row=row, column=column, value=cell_formula)

    wb.save(file_path)
    wb.close()


def calculate_statistics(file_path, columns, sheet_name=None):
    """
    计算统计信息（总和、平均、最大、最小）

    参数：
        file_path: str - Excel文件路径
        columns: list - 要统计的列号列表
        sheet_name: str - 工作表名称（可选）

    返回：
        dict - 统计结果

    示例调用：
        stats = calculate_statistics("成绩表.xlsx", [2, 3, 4])
        print(stats)
        # {'column_2': {'sum': 175, 'avg': 87.5, 'max': 90, 'min': 85}, ...}
    """
    if not OPENPYXL_AVAILABLE:
        return None

    headers, data = read_data(file_path, sheet_name)

    if not data:
        return None

    stats = {}

    for col in columns:
        values = [row[col-1] for row in data if isinstance(row[col-1], (int, float))]

        if values:
            col_name = headers[col-1] if headers and col <= len(headers) else f"列{col}"
            stats[col_name] = {
                'sum': sum(values),
                'avg': sum(values) / len(values),
                'max': max(values),
                'min': min(values),
                'count': len(values)
            }

    return stats


def create_grade_book_example():
    """
    示例：创建一个成绩表

    演示如何使用本工具创建一个完整的学生成绩表
    """
    if not OPENPYXL_AVAILABLE:
        print("❌ 请先安装 openpyxl: pip install openpyxl")
        return

    # 创建工作簿
    wb = create_workbook("成绩表")

    # 准备数据
    headers = ["姓名", "语文", "数学", "英语", "总分", "平均分", "等级"]
    students = [
        ["小明", 85, 92, 88],
        ["小红", 92, 88, 95],
        ["小刚", 78, 85, 82],
        ["小芳", 88, 90, 92],
        ["小华", 90, 95, 88]
    ]

    # 计算总分、平均分、等级
    for student in students:
        scores = student[1:]
        total = sum(scores)
        avg = total / len(scores)

        # 计算等级
        if avg >= 90:
            grade = "A"
        elif avg >= 80:
            grade = "B"
        elif avg >= 70:
            grade = "C"
        elif avg >= 60:
            grade = "D"
        else:
            grade = "F"

        student.extend([total, round(avg, 1), grade])

    # 写入数据
    all_data = [headers] + students
    write_data(wb, all_data)

    # 设置样式
    set_header_style(wb)
    auto_column_width(wb)

    # 设置所有数据居中
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存
    output_file = "test_files/成绩表示例.xlsx"
    os.makedirs("test_files", exist_ok=True)
    wb.save(output_file)

    print(f"✅ 成绩表已创建: {output_file}")

    # 显示统计信息
    stats = calculate_statistics(output_file, [2, 3, 4])
    print("\n📊 统计信息：")
    print("-" * 40)
    for subject, info in stats.items():
        print(f"{subject}:")
        print(f"  总分: {info['sum']}, 平均分: {info['avg']:.1f}")
        print(f"  最高分: {info['max']}, 最低分: {info['min']}")

    return output_file


def batch_process_excel(folder_path, operation, *args):
    """
    批量处理文件夹中的所有Excel文件

    参数：
        folder_path: str - 文件夹路径
        operation: str - 操作类型
        *args - 操作参数

    支持的操作：
        - "read": 读取所有Excel并合并数据
        - "stats": 计算所有Excel的统计信息

    示例调用：
        # 读取所有Excel文件
        all_data = batch_process_excel("data_folder", "read")
    """
    if not OPENPYXL_AVAILABLE:
        return None

    excel_files = [f for f in os.listdir(folder_path)
                   if f.endswith(('.xlsx', '.xls'))]

    if not excel_files:
        print(f"📁 文件夹中没有Excel文件: {folder_path}")
        return None

    results = []

    for filename in excel_files:
        file_path = os.path.join(folder_path, filename)

        if operation == "read":
            headers, data = read_data(file_path)
            results.append({
                'filename': filename,
                'headers': headers,
                'data': data
            })
            print(f"  📖 读取: {filename} ({len(data)}行数据)")

        elif operation == "stats":
            stats = calculate_statistics(file_path, args[0] if args else [1])
            results.append({
                'filename': filename,
                'stats': stats
            })

    return results


def print_menu():
    """打印菜单"""
    print("""
╔════════════════════════════════════════════════════════════╗
║               Excel处理工具 v1.0                           ║
╠════════════════════════════════════════════════════════════╣
║  1. 创建示例成绩表                                          ║
║  2. 读取Excel文件                                           ║
║  3. 计算统计信息                                            ║
║  4. 批量处理Excel文件                                       ║
║  q. 退出                                                    ║
╚════════════════════════════════════════════════════════════╝
    """)


def main():
    """主函数"""
    if not OPENPYXL_AVAILABLE:
        print("\n请先安装 openpyxl 库：")
        print("  pip install openpyxl")
        return

    while True:
        print_menu()
        choice = input("请选择操作: ").strip().lower()

        if choice == "q":
            print("👋 再见！")
            break

        elif choice == "1":
            create_grade_book_example()
            input("\n按回车继续...")

        elif choice == "2":
            file_path = input("请输入Excel文件路径: ").strip()
            headers, data = read_data(file_path)

            if headers and data:
                print(f"\n表头: {headers}")
                print(f"数据行数: {len(data)}")
                print("\n前5行数据:")
                for row in data[:5]:
                    print(f"  {row}")

            input("\n按回车继续...")

        elif choice == "3":
            file_path = input("请输入Excel文件路径: ").strip()
            cols = input("请输入要统计的列号（用逗号分隔，如 2,3,4）: ").strip()
            columns = [int(c.strip()) for c in cols.split(",")]

            stats = calculate_statistics(file_path, columns)

            if stats:
                print("\n📊 统计结果：")
                print("-" * 40)
                for col_name, info in stats.items():
                    print(f"\n{col_name}:")
                    print(f"  总和: {info['sum']}")
                    print(f"  平均: {info['avg']:.2f}")
                    print(f"  最大: {info['max']}")
                    print(f"  最小: {info['min']}")
                    print(f"  数量: {info['count']}")

            input("\n按回车继续...")

        elif choice == "4":
            folder_path = input("请输入文件夹路径: ").strip()
            results = batch_process_excel(folder_path, "read")

            if results:
                print(f"\n共读取 {len(results)} 个Excel文件")

            input("\n按回车继续...")

        else:
            print("❌ 无效选择，请重试\n")


if __name__ == "__main__":
    main()
